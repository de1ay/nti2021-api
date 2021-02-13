from rest_framework import viewsets
from rest_framework.decorators import action, permission_classes
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from openpyxl import load_workbook

from .serializers import *
from .models import *


class ItemViewSet(viewsets.ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer

    @action(detail=False, methods=['post'], serializer_class=FileSerializer)
    def upload_timetable(self, request):
        serializer = FileSerializer(data=request.data)
        if serializer.is_valid():
            wb = load_workbook(request.FILES['file'], data_only=True)
            sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
            for y in range(2, sheet.max_row):
                row = []
                for x in range(1, 8):
                    row.append(sheet.cell(row=y, column=x).value)
                try:
                    new_id = Item.objects.order_by('-address')[0].address + 1
                except Exception:
                    new_id = 0
                if new_id > 216:
                    return Response({"text": "На складе закончилось место"}, status=400)
                item = Item(name=row[0], model=row[1], system_id=row[2], date=row[3], time_in=row[4], time_out=row[5], vendor=row[6], address=new_id)
                item.save()
            return Response({"success": True})
        return Response(serializer.errors, status=400)


class LogViewSet(viewsets.ModelViewSet):
    queryset = Log.objects.all()
    serializer_class = LogSerializer

    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated], serializer_class=TextSerializer)
    def log_action(self, request):
        user = request.user
        serializer = TextSerializer(data=request.data)
        if serializer.is_valid():
            Log.objects.create(user=user, text=serializer.data['text'])
            return Response({"success": True})
        return Response(serializer.errors, status=400)
